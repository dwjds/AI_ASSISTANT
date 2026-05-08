from __future__ import annotations

import argparse
import copy
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Filter workbook rows by keyword criteria and write a new .xlsx file."
    )
    parser.add_argument("input", help="Input .xlsx/.xlsm file")
    parser.add_argument("output", help="Output .xlsx file")
    parser.add_argument("--sheet", help="Worksheet name; defaults to active sheet")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--data-start-row", type=int)
    parser.add_argument("--criteria-json", help="Criteria JSON string")
    parser.add_argument("--criteria-file", help="Path to criteria JSON file")
    parser.add_argument("--output-sheet", default="Filtered")
    parser.add_argument("--case-sensitive", action="store_true")
    args = parser.parse_args()

    try:
        payload = run(args)
    except Exception as exc:
        payload = {"status": "error", "error": str(exc)}
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if payload.get("status") == "success" else 1


def run(args: argparse.Namespace) -> dict[str, Any]:
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported.")

    criteria = load_criteria(args.criteria_json, args.criteria_file)
    include_groups = normalize_groups(criteria.get("include") or [])
    exclude_groups = normalize_groups(criteria.get("exclude") or [])
    if not include_groups and not exclude_groups:
        raise ValueError("At least one include or exclude criteria group is required.")

    keep_vba = input_path.suffix.lower() == ".xlsm"
    source_wb = load_workbook(str(input_path), keep_vba=keep_vba)
    try:
        source_sheet = source_wb[args.sheet] if args.sheet else source_wb.active
        header_row = max(1, int(args.header_row or 1))
        data_start_row = args.data_start_row or header_row + 1
        headers = read_headers(source_sheet, header_row)
        include_specs = resolve_groups(include_groups, headers)
        exclude_specs = resolve_groups(exclude_groups, headers)

        output_wb = Workbook()
        output_sheet = output_wb.active
        output_sheet.title = safe_sheet_name(args.output_sheet)
        copy_row(source_sheet, output_sheet, header_row, 1)

        matched_rows: list[int] = []
        for row_index in range(data_start_row, source_sheet.max_row + 1):
            if row_matches(source_sheet, row_index, include_specs, exclude_specs, args.case_sensitive):
                matched_rows.append(row_index)
                copy_row(source_sheet, output_sheet, row_index, output_sheet.max_row + 1)

        copy_dimensions(source_sheet, output_sheet)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_wb.save(str(output_path))
        output_wb.close()
        return {
            "status": "success",
            "input": str(input_path),
            "output": str(output_path),
            "source_sheet": source_sheet.title,
            "output_sheet": output_sheet.title,
            "header_row": header_row,
            "matched_count": len(matched_rows),
            "matched_source_rows": matched_rows[:100],
            "headers": headers,
            "criteria": criteria,
        }
    finally:
        source_wb.close()


def load_criteria(criteria_json: str | None, criteria_file: str | None) -> dict[str, Any]:
    if criteria_file:
        path = Path(criteria_file).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Criteria file not found: {path}")
        payload = path.read_text(encoding="utf-8-sig")
    else:
        payload = criteria_json or ""
    if not payload.strip():
        raise ValueError("--criteria-json or --criteria-file is required.")
    data = json.loads(payload)
    if not isinstance(data, dict):
        raise ValueError("Criteria must be a JSON object.")
    return data


def normalize_groups(groups: list[Any]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for group in groups:
        if not isinstance(group, dict):
            continue
        columns = [str(item).strip() for item in group.get("columns", []) if str(item).strip()]
        keywords = [str(item).strip() for item in group.get("keywords", []) if str(item).strip()]
        mode = str(group.get("mode") or "any").strip().lower()
        if columns and keywords:
            normalized.append({"columns": columns, "keywords": keywords, "mode": "all" if mode == "all" else "any"})
    return normalized


def read_headers(sheet: Any, header_row: int) -> list[str]:
    headers: list[str] = []
    for col_index in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col_index).value
        headers.append(str(value or "").strip())
    return headers


def resolve_groups(groups: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    resolved: list[dict[str, Any]] = []
    for group in groups:
        column_indexes = [resolve_column(column, headers) for column in group["columns"]]
        resolved.append({**group, "column_indexes": column_indexes})
    return resolved


def resolve_column(column: str, headers: list[str]) -> int:
    text = str(column).strip()
    if text.isdigit():
        index = int(text)
        if index < 1 or index > len(headers):
            raise ValueError(f"Column index out of range: {column}")
        return index
    for index, header in enumerate(headers, start=1):
        if header.lower() == text.lower():
            return index
    raise ValueError(f"Column/header not found: {column}. Available headers: {headers}")


def row_matches(
    sheet: Any,
    row_index: int,
    include_specs: list[dict[str, Any]],
    exclude_specs: list[dict[str, Any]],
    case_sensitive: bool,
) -> bool:
    for spec in include_specs:
        if not group_matches(sheet, row_index, spec, case_sensitive):
            return False
    for spec in exclude_specs:
        if group_matches(sheet, row_index, spec, case_sensitive):
            return False
    return True


def group_matches(sheet: Any, row_index: int, spec: dict[str, Any], case_sensitive: bool) -> bool:
    haystack_parts = [
        str(sheet.cell(row=row_index, column=col_index).value or "")
        for col_index in spec["column_indexes"]
    ]
    haystack = " ".join(haystack_parts)
    keywords = list(spec["keywords"])
    if not case_sensitive:
        haystack = haystack.lower()
        keywords = [keyword.lower() for keyword in keywords]
    hits = [keyword in haystack for keyword in keywords]
    return all(hits) if spec.get("mode") == "all" else any(hits)


def copy_row(source_sheet: Any, output_sheet: Any, source_row: int, target_row: int) -> None:
    output_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
    for col_index in range(1, source_sheet.max_column + 1):
        source_cell = source_sheet.cell(row=source_row, column=col_index)
        target_cell = output_sheet.cell(row=target_row, column=col_index, value=source_cell.value)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)


def copy_dimensions(source_sheet: Any, output_sheet: Any) -> None:
    for key, dimension in source_sheet.column_dimensions.items():
        output_sheet.column_dimensions[key].width = dimension.width
        output_sheet.column_dimensions[key].hidden = dimension.hidden


def safe_sheet_name(value: str) -> str:
    cleaned = str(value or "Filtered").strip() or "Filtered"
    for char in "[]:*?/\\": 
        cleaned = cleaned.replace(char, "_")
    return cleaned[:31]


if __name__ == "__main__":
    raise SystemExit(main())
