"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import os
import platform
import subprocess
import sys
import tempfile
from pathlib import Path

from office.soffice import get_soffice_env

from openpyxl import load_workbook

MACRO_FILENAME = "Module1.xba"
PROFILE_DIRNAME = "miniagent_lo_profile_xlsx"

SCRIPT_XLC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink">
 <library:library library:name="Standard" xlink:href="$(USER)/basic/Standard/script.xlb/" xlink:type="simple" library:link="false"/>
</library:libraries>"""

DIALOG_XLC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink">
 <library:library library:name="Standard" xlink:href="$(USER)/basic/Standard/dialog.xlb/" xlink:type="simple" library:link="false"/>
</library:libraries>"""

SCRIPT_XLB = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="Module1"/>
</library:library>"""

DIALOG_XLB = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false"/>"""

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def get_profile_dir() -> Path:
    base_dir = Path(tempfile.gettempdir()) / PROFILE_DIRNAME
    if os.access(str(base_dir.parent), os.W_OK):
        return base_dir
    fallback = Path(__file__).resolve().parent / ".runtime" / PROFILE_DIRNAME
    return fallback


def get_profile_uri() -> str:
    return get_profile_dir().resolve().as_uri()


def get_macro_dir() -> Path:
    return get_profile_dir() / "user" / "basic" / "Standard"


def _initialize_profile() -> tuple[bool, str]:
    profile_dir = get_profile_dir()
    profile_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        "soffice",
        "--headless",
        f"-env:UserInstallation={get_profile_uri()}",
        "--terminate_after_init",
    ]
    try:
        subprocess.run(
            cmd,
            capture_output=True,
            timeout=10,
            env=get_soffice_env(),
            check=False,
        )
    except FileNotFoundError:
        return False, "LibreOffice soffice executable was not found."
    except Exception as exc:
        return False, str(exc)
    return True, ""


def _write_macro_library_files(macro_dir: Path) -> tuple[bool, str]:
    try:
        macro_dir.mkdir(parents=True, exist_ok=True)
        basic_dir = macro_dir.parent
        (basic_dir / "script.xlc").write_text(SCRIPT_XLC, encoding="utf-8")
        (basic_dir / "dialog.xlc").write_text(DIALOG_XLC, encoding="utf-8")
        (macro_dir / "script.xlb").write_text(SCRIPT_XLB, encoding="utf-8")
        (macro_dir / "dialog.xlb").write_text(DIALOG_XLB, encoding="utf-8")
        (macro_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True, ""
    except Exception as exc:
        return False, str(exc)


def setup_libreoffice_macro():
    macro_dir = get_macro_dir()
    macro_file = macro_dir / MACRO_FILENAME

    try:
        if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text(encoding="utf-8"):
            return True, ""
    except Exception as exc:
        return False, f"Failed to inspect existing macro file: {exc}"

    initialized, init_error = _initialize_profile()
    if not initialized:
        return False, init_error

    written, write_error = _write_macro_library_files(macro_dir)
    if not written:
        return False, write_error
    return True, ""


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    ready, macro_error = setup_libreoffice_macro()
    if not ready:
        return {"error": f"Failed to setup LibreOffice macro: {macro_error}"}

    cmd = [
        "soffice",
        "--headless",
        f"-env:UserInstallation={get_profile_uri()}",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            env=get_soffice_env(),
            timeout=timeout if platform.system() == "Windows" else None,
        )
    except FileNotFoundError:
        return {"error": "LibreOffice soffice executable was not found. Install LibreOffice and ensure soffice is on PATH."}
    except subprocess.TimeoutExpired:
        result = subprocess.CompletedProcess(cmd, 124, "", "")

    if result.returncode != 0 and result.returncode != 124:  
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],  
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
