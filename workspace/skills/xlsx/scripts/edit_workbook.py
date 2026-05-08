from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Apply common Excel workbook edits using openpyxl and return JSON."
    )
    parser.add_argument("input", help="Input .xlsx/.xlsm file")
    parser.add_argument("output", help="Output workbook path")
    parser.add_argument(
        "--operation",
        required=True,
        choices=[
            "inspect",
            "read-range",
            "set-cell",
            "append-row",
            "insert-row",
            "delete-row",
            "insert-column",
            "delete-column",
            "set-header",
            "add-formula-column",
            "add-sum-row",
            "rename-sheet",
            "copy-sheet",
            "delete-sheet",
        ],
    )
    parser.add_argument("--sheet", help="Worksheet name; defaults to active sheet")
    parser.add_argument("--range", dest="cell_range", help="A1 range for read-range")
    parser.add_argument("--cell", help="A1 cell coordinate for set-cell")
    parser.add_argument("--value", help="Cell/header/formula/label value")
    parser.add_argument("--values", nargs="*", help="Row values for append/insert")
    parser.add_argument("--row", type=int, help="1-based row index")
    parser.add_argument("--column", help="Column letter, number, or header name")
    parser.add_argument("--columns", nargs="*", help="Column letters/numbers/header names")
    parser.add_argument("--label-column", help="Column letter/number/header for row label")
    parser.add_argument("--label", default="Total", help="Label used by add-sum-row")
    parser.add_argument("--new-name", help="New sheet name for rename/copy")
    parser.add_argument("--formula", help="Formula template for add-formula-column")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--data-start-row", type=int)
    parser.add_argument("--data-end-row", type=int)
    parser.add_argument("--preserve-formulas", action="store_true", default=True)
    args = parser.parse_args()

    try:
        payload = run(args)
    except Exception as exc:
        payload = {"status": "error", "error": str(exc), "operation": args.operation}
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if payload.get("status") == "success" else 1


def run(args: argparse.Namespace) -> dict[str, Any]:
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported.")

    keep_vba = input_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(str(input_path), keep_vba=keep_vba)
    try:
        sheet = get_sheet(workbook, args.sheet)
        operation = args.operation
        result: dict[str, Any]

        if operation == "inspect":
            result = inspect_workbook(workbook)
        elif operation == "read-range":
            result = read_range(sheet, args.cell_range)
        elif operation == "set-cell":
            result = set_cell(sheet, args.cell, parse_value(args.value))
        elif operation == "append-row":
            result = append_row(sheet, args.values or [])
        elif operation == "insert-row":
            result = insert_row(sheet, args.row, args.values or [])
        elif operation == "delete-row":
            result = delete_row(sheet, args.row)
        elif operation == "insert-column":
            result = insert_column(sheet, args.column, args.value)
        elif operation == "delete-column":
            result = delete_column(sheet, args.column, args.header_row)
        elif operation == "set-header":
            result = set_header(sheet, args.column, args.value, args.header_row)
        elif operation == "add-formula-column":
            result = add_formula_column(
                sheet,
                args.column,
                args.value,
                args.formula,
                args.header_row,
                args.data_start_row,
                args.data_end_row,
            )
        elif operation == "add-sum-row":
            result = add_sum_row(
                sheet,
                args.columns or [],
                args.label_column,
                args.label,
                args.header_row,
                args.data_start_row,
                args.data_end_row,
            )
        elif operation == "rename-sheet":
            result = rename_sheet(sheet, args.new_name)
        elif operation == "copy-sheet":
            result = copy_sheet(workbook, sheet, args.new_name)
        elif operation == "delete-sheet":
            result = delete_sheet(workbook, sheet)
        else:
            raise ValueError(f"Unsupported operation: {operation}")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        result.update(
            {
                "status": "success",
                "operation": operation,
                "input": str(input_path),
                "output": str(output_path),
            }
        )
        return result
    finally:
        workbook.close()


def get_sheet(workbook: Any, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {workbook.sheetnames}")
        return workbook[sheet_name]
    return workbook.active


def inspect_workbook(workbook: Any) -> dict[str, Any]:
    sheets = []
    for sheet in workbook.worksheets:
        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, min(sheet.max_column, 20) + 1)
        ]
        sheets.append(
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "headers_preview": [value for value in headers if value is not None],
            }
        )
    return {"sheets": sheets}


def read_range(sheet: Any, cell_range: str | None) -> dict[str, Any]:
    if not cell_range:
        cell_range = f"A1:{get_column_letter(min(sheet.max_column, 10))}{min(sheet.max_row, 20)}"
    rows: list[list[Any]] = []
    for row in sheet[cell_range]:
        rows.append([cell.value for cell in row])
    return {"sheet": sheet.title, "range": cell_range, "values": rows}


def set_cell(sheet: Any, cell: str | None, value: Any) -> dict[str, Any]:
    if not cell:
        raise ValueError("--cell is required for set-cell.")
    sheet[cell] = value
    return {"sheet": sheet.title, "cell": cell, "value": value}


def append_row(sheet: Any, values: list[str]) -> dict[str, Any]:
    parsed = [parse_value(value) for value in values]
    sheet.append(parsed)
    return {"sheet": sheet.title, "row": sheet.max_row, "values": parsed}


def insert_row(sheet: Any, row_index: int | None, values: list[str]) -> dict[str, Any]:
    if not row_index or row_index < 1:
        raise ValueError("--row must be a positive integer.")
    sheet.insert_rows(row_index)
    parsed = [parse_value(value) for value in values]
    for col_index, value in enumerate(parsed, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)
    return {"sheet": sheet.title, "row": row_index, "values": parsed}


def delete_row(sheet: Any, row_index: int | None) -> dict[str, Any]:
    if not row_index or row_index < 1:
        raise ValueError("--row must be a positive integer.")
    sheet.delete_rows(row_index)
    return {"sheet": sheet.title, "deleted_row": row_index}


def insert_column(sheet: Any, column: str | None, header: str | None) -> dict[str, Any]:
    col_index = resolve_column(sheet, column, header_row=1, allow_next=True)
    sheet.insert_cols(col_index)
    if header:
        sheet.cell(row=1, column=col_index, value=header)
    return {"sheet": sheet.title, "inserted_column": get_column_letter(col_index), "header": header}


def delete_column(sheet: Any, column: str | None, header_row: int) -> dict[str, Any]:
    col_index = resolve_column(sheet, column, header_row=header_row)
    sheet.delete_cols(col_index)
    return {"sheet": sheet.title, "deleted_column": get_column_letter(col_index)}


def set_header(sheet: Any, column: str | None, header: str | None, header_row: int) -> dict[str, Any]:
    if not header:
        raise ValueError("--value is required for set-header.")
    col_index = resolve_column(sheet, column, header_row=header_row, allow_next=True)
    sheet.cell(row=header_row, column=col_index, value=header)
    return {"sheet": sheet.title, "column": get_column_letter(col_index), "header": header}


def add_formula_column(
    sheet: Any,
    column: str | None,
    header: str | None,
    formula_template: str | None,
    header_row: int,
    data_start_row: int | None,
    data_end_row: int | None,
) -> dict[str, Any]:
    if not header:
        raise ValueError("--value is required as the new column header.")
    if not formula_template:
        raise ValueError("--formula is required for add-formula-column.")
    col_index = resolve_column(sheet, column, header_row=header_row, allow_next=True)
    if col_index <= sheet.max_column:
        sheet.insert_cols(col_index)
    sheet.cell(row=header_row, column=col_index, value=header)
    start_row = data_start_row or header_row + 1
    end_row = data_end_row or sheet.max_row
    formulas = []
    for row in range(start_row, end_row + 1):
        formula = build_formula(formula_template, row)
        sheet.cell(row=row, column=col_index, value=formula)
        formulas.append({"cell": f"{get_column_letter(col_index)}{row}", "formula": formula})
    return {
        "sheet": sheet.title,
        "column": get_column_letter(col_index),
        "header": header,
        "formula_count": len(formulas),
        "formulas_preview": formulas[:10],
    }


def add_sum_row(
    sheet: Any,
    columns: list[str],
    label_column: str | None,
    label: str,
    header_row: int,
    data_start_row: int | None,
    data_end_row: int | None,
) -> dict[str, Any]:
    if not columns:
        raise ValueError("--columns is required for add-sum-row.")
    start_row = data_start_row or header_row + 1
    end_row = data_end_row or sheet.max_row
    target_row = sheet.max_row + 1
    label_col_index = resolve_column(sheet, label_column, header_row=header_row) if label_column else 1
    sheet.cell(row=target_row, column=label_col_index, value=label)
    formulas = []
    for column in columns:
        col_index = resolve_column(sheet, column, header_row=header_row)
        letter = get_column_letter(col_index)
        formula = f"=SUM({letter}{start_row}:{letter}{end_row})"
        sheet.cell(row=target_row, column=col_index, value=formula)
        formulas.append({"cell": f"{letter}{target_row}", "formula": formula})
    copy_row_style(sheet, end_row, target_row)
    return {
        "sheet": sheet.title,
        "sum_row": target_row,
        "label": label,
        "columns": columns,
        "formulas": formulas,
    }


def rename_sheet(sheet: Any, new_name: str | None) -> dict[str, Any]:
    if not new_name:
        raise ValueError("--new-name is required for rename-sheet.")
    old_name = sheet.title
    sheet.title = new_name[:31]
    return {"old_name": old_name, "new_name": sheet.title}


def copy_sheet(workbook: Any, sheet: Any, new_name: str | None) -> dict[str, Any]:
    copied = workbook.copy_worksheet(sheet)
    copied.title = (new_name or f"{sheet.title}_copy")[:31]
    return {"source_sheet": sheet.title, "new_sheet": copied.title}


def delete_sheet(workbook: Any, sheet: Any) -> dict[str, Any]:
    if len(workbook.worksheets) <= 1:
        raise ValueError("Cannot delete the only worksheet in a workbook.")
    title = sheet.title
    workbook.remove(sheet)
    return {"deleted_sheet": title}


def resolve_column(sheet: Any, column: str | None, *, header_row: int, allow_next: bool = False) -> int:
    if not column:
        if allow_next:
            return sheet.max_column + 1
        raise ValueError("--column is required.")
    text = str(column).strip()
    if not text:
        if allow_next:
            return sheet.max_column + 1
        raise ValueError("--column is required.")
    if text.isdigit():
        return int(text)
    if re.fullmatch(r"[A-Za-z]{1,3}", text):
        value = 0
        for char in text.upper():
            value = value * 26 + (ord(char) - ord("A") + 1)
        return value
    for col_index in range(1, sheet.max_column + 1):
        header = sheet.cell(row=header_row, column=col_index).value
        if str(header or "").strip().lower() == text.lower():
            return col_index
    if allow_next:
        return sheet.max_column + 1
    raise ValueError(f"Column/header not found: {column}")


def parse_value(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    stripped = text.strip()
    if stripped == "":
        return ""
    if stripped.startswith("="):
        return stripped
    lowered = stripped.lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    if lowered in {"none", "null"}:
        return None
    try:
        if "." not in stripped:
            return int(stripped)
        return float(stripped)
    except ValueError:
        return text


def build_formula(template: str, row: int) -> str:
    formula = template.replace("{row}", str(row)).replace("{{row}}", str(row))
    return formula if formula.startswith("=") else f"={formula}"


def copy_row_style(sheet: Any, source_row: int, target_row: int):
    if source_row < 1 or target_row < 1:
        return
    for col_index in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=col_index)
        target = sheet.cell(row=target_row, column=col_index)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)


if __name__ == "__main__":
    raise SystemExit(main())
