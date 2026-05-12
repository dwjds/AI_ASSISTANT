from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:  # pragma: no cover
    colors = None  # type: ignore[assignment]
    A4 = None  # type: ignore[assignment]
    landscape = None  # type: ignore[assignment]
    getSampleStyleSheet = None  # type: ignore[assignment]
    pdfmetrics = None  # type: ignore[assignment]
    UnicodeCIDFont = None  # type: ignore[assignment]
    Paragraph = None  # type: ignore[assignment]
    SimpleDocTemplate = None  # type: ignore[assignment]
    Spacer = None  # type: ignore[assignment]
    Table = None  # type: ignore[assignment]
    TableStyle = None  # type: ignore[assignment]


def main() -> int:
    parser = argparse.ArgumentParser(description="Export an XLSX/XLSM workbook sheet to a simple PDF table.")
    parser.add_argument("input", help="Input .xlsx/.xlsm file")
    parser.add_argument("output", help="Output .pdf file")
    parser.add_argument("--sheet", default="", help="Optional sheet name. Defaults to the active sheet.")
    parser.add_argument("--title", default="", help="Optional PDF title.")
    parser.add_argument("--max-rows", type=int, default=200, help="Maximum data rows to export.")
    args = parser.parse_args()

    try:
        payload = convert_workbook_to_pdf(
            input_path=Path(args.input),
            output_path=Path(args.output),
            sheet_name=args.sheet,
            title=args.title,
            max_rows=args.max_rows,
        )
    except Exception as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def convert_workbook_to_pdf(
    *,
    input_path: Path,
    output_path: Path,
    sheet_name: str = "",
    title: str = "",
    max_rows: int = 200,
) -> dict[str, Any]:
    source = input_path.expanduser().resolve()
    target = output_path.expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"Input workbook not found: {source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported.")
    if target.suffix.lower() != ".pdf":
        raise ValueError("Output path must end with .pdf.")

    libreoffice_payload = try_convert_with_libreoffice(source, target)
    if libreoffice_payload is not None:
        return libreoffice_payload

    if SimpleDocTemplate is None:
        raise RuntimeError("Exporting PDF requires LibreOffice or reportlab.")

    workbook = load_workbook(filename=str(source), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows: list[list[str]] = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = ["" if value is None else str(value) for value in row]
        if any(value.strip() for value in values):
            rows.append(values)
        if len(rows) >= max_rows:
            break
    if not rows:
        rows = [["No data"]]

    width = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in rows]

    target.parent.mkdir(parents=True, exist_ok=True)
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = "STSong-Light"

    doc = SimpleDocTemplate(
        str(target),
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    story: list[Any] = []
    heading = title or worksheet.title
    story.append(Paragraph(heading, styles["Title"]))
    story.append(Spacer(1, 12))

    table_data = [[Paragraph(cell.replace("\n", "<br/>"), styles["BodyText"]) for cell in row] for row in normalized_rows]
    col_widths = [doc.width / max(width, 1)] * width
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEFF7")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)

    return {
        "status": "success",
        "input": str(source),
        "output": str(target),
        "sheet": worksheet.title,
        "rows_exported": len(normalized_rows),
        "columns_exported": width,
        "engine": "reportlab",
    }


def try_convert_with_libreoffice(source: Path, target: Path) -> dict[str, Any] | None:
    soffice = shutil.which("soffice")
    if not soffice:
        return None
    soffice_path = Path(soffice)
    if soffice_path.suffix.lower() == ".com":
        exe_candidate = soffice_path.with_suffix(".exe")
        if exe_candidate.exists():
            soffice = str(exe_candidate)
    target.parent.mkdir(parents=True, exist_ok=True)
    generated = target.parent / f"{source.stem}.pdf"
    if generated.exists() and generated.resolve() != target.resolve():
        generated.unlink()
    profile_dir = Path(tempfile.mkdtemp(prefix="miniagent_lo_profile_")).resolve()
    profile_uri = "file:///" + str(profile_dir).replace("\\", "/")
    command = [
        soffice,
        "--headless",
        f"-env:UserInstallation={profile_uri}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(target.parent),
        str(source),
    ]
    proc = subprocess.run(command, capture_output=True, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError(
            "LibreOffice conversion failed: "
            + (proc.stderr.strip() or proc.stdout.strip() or f"return code {proc.returncode}")
        )
    if not generated.exists():
        candidates = sorted(target.parent.glob("*.pdf"), key=lambda path: path.stat().st_mtime, reverse=True)
        generated = candidates[0] if candidates else generated
    if not generated.exists():
        raise RuntimeError("LibreOffice conversion finished but no PDF was produced.")
    if generated.resolve() != target.resolve():
        if target.exists():
            target.unlink()
        generated.rename(target)
    return {
        "status": "success",
        "input": str(source),
        "output": str(target),
        "engine": "libreoffice",
        "stdout": proc.stdout.strip(),
    }


if __name__ == "__main__":
    raise SystemExit(main())
